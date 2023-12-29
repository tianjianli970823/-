import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook

# 定义数据库连接信息
db_user = 'root'
db_password = '970823'
db_host = 'localhost'
db_name = 'an4006'

# 创建SQLAlchemy引擎并连接到MySQL数据库
engine = create_engine(f"mysql+pymysql://{db_user}:{db_password}@{db_host}/{db_name}?charset=utf8mb4")

# 从数据库中读取订单数量大于0的数据
df = pd.read_sql_query('SELECT * FROM an1004 WHERE 订单数量 > 0', con=engine)

# 加载Excel文件
book = load_workbook(r"D:\数据源\Test1\导入配对商品模板 (按MSKU).xlsx")

# 选择要修改的工作表
sheet = book['Sheet1']

# 创建一个字典，键是Excel中的列名，值是数据库中的列名
columns_dict = {
    '*SKU': 'SKU',
    '*MSKU': 'SKU',
    '店铺': '店铺',
}
# 对于字典中的每一对键值，获取Excel中的列位置，然后将数据库中的数据写入到Excel中的对应列
for excel_column, db_column in columns_dict.items():
    # 获取列的位置
    column_index = [cell.column for cell in sheet[1] if cell.value == excel_column][0]

    # 获取数据库中的数据
    data = df[db_column].tolist()

    # 将数据写入到Excel中的对应列
    for i in range(len(data)):
        sheet.cell(row=i + 2, column=column_index, value=data[i])

# 获取 "是否同步listing图" 列的位置
同步图_column = [cell for cell in sheet[1] if cell.value == '是否同步listing图'][0].column

# 在 "是否同步listing图" 列的每一行写入 "否"
for i in range(2, len(df) + 2):
    sheet.cell(row=i, column=同步图_column, value='否')


# 保存Excel文件
book.save(r"D:\数据源\Test1\导入配对商品模板 (按MSKU).xlsx")