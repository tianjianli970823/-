#根据FNSKU匹配SKU和名字，从数据库导入到excel中 part2
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook

# 定义数据库连接信息
db_user = 'root'
db_password = '970823'
db_host = 'localhost'
db_name = 'an4006'

# 创建SQLAlchemy引擎并连接到MySQL数据库
engine = create_engine(f"mysql+pymysql://{db_user}:{db_password}@{db_host}/{db_name}?charset=utf8mb4")

# 从数据库中读取订单数量大于0的数据
df = pd.read_sql_query('SELECT * FROM an1014 WHERE 订单数量 > 0', con=engine)

# 加载Excel文件
book = load_workbook(r"D:\数据源\PO下单表-数据源\12.25-AN-空派.xlsx")

# 选择要修改的工作表
sheet = book['基本资料 ']

# 创建一个字典，键是Excel中的列名，值是数据库中的列名
columns_dict = {
    'FNSKU': 'SKU',
    '商品编码': 'SKU',
    '品名': '颜色',
}

# 遍历字典
for excel_column, db_column in columns_dict.items():
    # 找到Excel中列名对应的列
    for cell in sheet[1]:
        if cell.value == excel_column:
            excel_col_index = cell.col_idx
            break

    # 遍历数据库查询结果的每一行
    for index, row in df.iterrows():
        # 在Excel中找到对应的行
        excel_row_index = index + 2  # +2 是因为Excel的行索引从1开始，且第一行通常是标题行

        # 从数据库查询结果中获取对应的值
        value = row[db_column]

        # 更新Excel中的值
        sheet.cell(row=excel_row_index, column=excel_col_index, value=value)



# 保存Excel文件
book.save(r"D:\数据源\PO下单表-数据源\12.25-AN-空派.xlsx")


