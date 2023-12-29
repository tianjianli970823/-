import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook

# 定义数据库连接信息
db_user = 'root'
db_password = '970823'
db_host = 'localhost'
db_name = 'an4006'

# 创建SQLAlchemy引擎并连接到MySQL数据库
engine = create_engine(f"mysql+pymysql://{db_user}:{db_password}@{db_host}/{db_name}?charset=utf8mb4")

# 从数据库中读取订单数量大于0的数据
df = pd.read_sql_query('SELECT * FROM an1004 WHERE 订单数量 > 0', con=engine)

# 加载Excel文件
book = load_workbook(r"D:\数据源\Test1\领星产品批量导入.xlsx")

# 选择要修改的工作表
sheet = book['产品导入模板']

# 创建一个字典，键是Excel中的列名，值是数据库中的列名
columns_dict = {
    '*SKU': 'SKU',
    '*品名': '款号',
    '型号': '型号',
    '一级分类': '一级分类',
    '二级分类': '二级分类',
    '品牌': 'brand_name',
    '产品负责人': '产品负责人'
}

# 对于字典中的每一对键值，获取Excel中的列位置，然后将数据库中的数据写入到Excel中的对应列
for excel_column, db_column in columns_dict.items():
    # 获取列的位置
    column_index = [cell.column for cell in sheet[1] if cell.value == excel_column][0]

    # 获取数据库中的数据
    data = df[db_column].tolist()

    # 将数据写入到Excel中的对应列
    for i in range(len(data)):
        sheet.cell(row=i + 2, column=column_index, value=data[i])




# 保存Excel文件
book.save(r"D:\数据源\Test1\领星产品批量导入.xlsx")


