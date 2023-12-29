#根据FNSKU匹配SKU和名字，从数据库导入到excel中
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook

# 定义数据库连接信息
db_user = 'root'
db_password = '970823'
db_host = 'localhost'
db_name = 'an4006'

# 创建SQLAlchemy引擎并连接到MySQL数据库
engine = create_engine(f"mysql+pymysql://{db_user}:{db_password}@{db_host}/{db_name}?charset=utf8mb4")

# 从数据库中读取订单数量大于0的数据
df = pd.read_sql_query('SELECT * FROM an1014 WHERE 订单数量 > 0', con=engine)

# 加载Excel文件
book = load_workbook(r"D:\数据源\PO下单表-数据源\12.25-AN-空派.xlsx")

# 选择要修改的工作表
sheet = book['工作表']

# 创建一个字典，键是Excel中的列名，值是数据库中的列名
columns_dict = {
    'FNSKU（扫码）': 'FNSKU',
    'SKU（抓取）': 'SKU',
    '名称（抓取）': '颜色',
}

# 创建一个临时字典来存储'FNSKU（扫码）'列的值和对应的'SKU（抓取）'和'名称（抓取）'列的值
temp_dict = {}

# 遍历数据库查询结果的每一行
for index, row in df.iterrows():
    # 从数据库查询结果中获取对应的值
    fnsku_value = row[columns_dict['FNSKU（扫码）']]
    sku_value = row[columns_dict['SKU（抓取）']]
    color_value = row[columns_dict['名称（抓取）']]

    # 将'FNSKU（扫码）'列的值和对应的'SKU（抓取）'和'名称（抓取）'列的值存储到临时字典中
    temp_dict[fnsku_value] = (sku_value, color_value)

# 获取Excel文件中的最大行数
max_row = sheet.max_row

# 获取Excel中的列标题行
header_row = sheet[1]

# 创建一个字典，键是Excel中的列名，值是列索引
excel_columns_dict = {cell.value: cell.column for cell in header_row}

# 遍历Excel中的每一行
for i in range(2, max_row + 1):  # 从第二行开始遍历，因为第一行是标题行
    # 获取'FNSKU（扫码）'列的值
    fnsku_value = sheet.cell(row=i, column=excel_columns_dict['FNSKU（扫码）']).value

    # 如果'FNSKU（扫码）'列的值在临时字典中存在
    if fnsku_value in temp_dict:
        # 获取对应的'SKU（抓取）'和'名称（抓取）'列的值
        sku_value, color_value = temp_dict[fnsku_value]

        # 更新Excel中的值
        sheet.cell(row=i, column=excel_columns_dict['SKU（抓取）'], value=sku_value)
        sheet.cell(row=i, column=excel_columns_dict['名称（抓取）'], value=color_value)

# 保存Excel文件
book.save(r"D:\数据源\PO下单表-数据源\12.25-AN-空派.xlsx")
